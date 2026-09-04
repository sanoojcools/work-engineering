"""Builds every file in offer-desk-inputs/ from the shared candidate roster
in candidates.py. Run once: python3 build_all.py
Everything produced is FABRICATED FOR TESTING -- see the top-level README.
"""
import csv
import json
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from candidates import CANDIDATES

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

HEADER_FILL = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 45)


def write_xlsx(path, headers, rows, sheet_title="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    autosize(ws)
    full = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    wb.save(full)
    print("wrote", path)


def write_csv(path, headers, rows):
    full = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    print("wrote", path)


def write_json(path, data):
    full = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w") as f:
        json.dump(data, f, indent=2)
    print("wrote", path)


def dplus(date_str, days):
    return (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=days)).strftime("%Y-%m-%d")


# ---------------------------------------------------------------- 00 -------
def build_candidate_roster():
    headers = [
        "Candidate ID", "Application ID", "Name", "Hire Type", "Function",
        "Grade", "City", "DOJ", "Recruiter", "Hiring Manager", "CTC (INR)",
        "Variable Pay %", "Salary Deviation %", "Joining Bonus (INR)",
        "Deferred Bonus (INR)", "Deferred Bonus Vest Date", "ESOP",
        "Source Type", "Source Name", "UAN Dual-Employment Flag", "Notes",
    ]
    rows = []
    for c in CANDIDATES:
        rows.append([
            c["id"], c["application_id"], c["name"], c["hire_type"], c["function"],
            c["grade"], c["city"], c["doj"], c["recruiter"], c["manager"],
            c["ctc"] or c.get("contract_amount") or c.get("stipend") or "",
            c["variable_pct"], c["deviation_pct"], c["joining_bonus"],
            c["deferred_bonus"], c["deferred_bonus_vest"], "Yes" if c["esop"] else "No",
            c["source_type"], c["source_name"], "Yes" if c["uan_overlap"] else "No",
            c["flag"],
        ])
    write_xlsx("offer-desk-inputs/00-foundation/candidate-roster.xlsx", headers, rows, "Roster")


# ---------------------------------------------------------- 03 Zwayam ------
def build_zwayam_export():
    events = []
    doc_types_permanent = ["EF", "MJS", "10th_certificate", "graduation_certificate",
                            "service_letters", "UAN_service_history", "EPFO_history",
                            "offer_letter_prev", "payslip_1", "payslip_2", "payslip_3",
                            "aadhaar", "pan", "passport_photo", "IEF_screenshot", "resume"]
    doc_types_contractor = ["EF", "MJS_contractor", "deputation_letter", "aadhaar", "pan",
                             "passport_photo", "IEF_screenshot"]
    doc_types_fresher = ["EF", "MJS", "10th_certificate", "graduation_certificate",
                          "aadhaar", "pan", "passport_photo", "resume"]
    doc_types_intern = ["EF", "aadhaar", "pan", "passport_photo"]

    for c in CANDIDATES:
        base = dplus(c["doj"], -28)
        events.append({
            "event": "candidate_profile_created", "candidate_id": c["id"],
            "application_id": c["application_id"], "name": c["name"],
            "hire_type": c["hire_type"], "function": c["function"], "grade": c["grade"],
            "timestamp": f"{base}T09:14:00+05:30",
        })
        docs = {"permanent": doc_types_permanent, "contractor": doc_types_contractor,
                "campus": doc_types_fresher, "fresher": doc_types_fresher,
                "intern": doc_types_intern}[c["hire_type"]]
        # C-2026-0146 (Divya) deliberately missing one payslip + flagged UAN overlap
        for i, doc in enumerate(docs):
            if c["id"] == "C-2026-0146" and doc == "payslip_3":
                continue  # missing document, matches step-2 "chasing missing documents" pain
            events.append({
                "event": "document_uploaded", "candidate_id": c["id"],
                "doc_type": doc, "timestamp": f"{dplus(base, i)}T{10 + (i % 6)}:{(i*7)%60:02d}:00+05:30",
                "verification_status": "pending",
            })
        verified_ts = dplus(base, len(docs) + 2)
        events.append({
            "event": "document_verification_completed", "candidate_id": c["id"],
            "timestamp": f"{verified_ts}T16:40:00+05:30",
            "result": "dual_employment_stop" if c["uan_overlap"] else (
                "missing_documents" if c["id"] == "C-2026-0146" else "complete"),
            "verified_by": "Rashmi KN",
        })
        if c["deviation_pct"]:
            events.append({
                "event": "salary_deviation_approval_triggered", "candidate_id": c["id"],
                "deviation_pct": c["deviation_pct"], "approvers": ["Vasu", "Nagaraj"],
                "timestamp": f"{dplus(verified_ts, 1)}T11:00:00+05:30", "status": "approved",
            })
        if not c["uan_overlap"]:
            events.append({
                "event": "offer_letter_trigger", "candidate_id": c["id"],
                "timestamp": f"{dplus(verified_ts, 2)}T09:30:00+05:30",
                "pushed_to": "Zoho",
            })
    write_json("offer-desk-inputs/03-system-logs/zwayam-candidate-export.json", {
        "_disclaimer": "FABRICATED FOR TESTING. Not a real Zwayam export.",
        "export_generated": "2026-07-02T08:00:00+05:30",
        "source_system": "Zwayam (synthetic)",
        "event_count": len(events),
        "events": events,
    })


# ------------------------------------------------------------ 03 Zoho ------
def build_zoho_log():
    rows = []
    for c in CANDIDATES:
        if c["uan_overlap"]:
            continue  # offer never released -> no Zoho event
        sent = dplus(c["doj"], -18)
        signed_by_th = dplus(sent, 1)
        signed_by_cand = dplus(signed_by_th, 2 if c["hire_type"] != "campus" else 4)
        rows.append([c["id"], c["name"], "offer_letter", sent, "Nagaraj T (TA Head)",
                     signed_by_th, "candidate", signed_by_cand, "signed"])
        ack_sent = dplus(c["doj"], -3)
        rows.append([c["id"], c["name"], "onboarding_acknowledgement", ack_sent,
                     "candidate", dplus(ack_sent, 1),
                     f"{c['city']} location SPOC", dplus(ack_sent, 1), "signed"])
    write_csv("offer-desk-inputs/03-system-logs/zoho-signing-log.csv",
              ["candidate_id", "name", "document_type", "sent_date", "signer_1",
               "signer_1_signed_date", "signer_2", "signer_2_signed_date", "status"],
              rows)


# ------------------------------------------------------------- 03 UAN ------
def build_uan_history():
    rows = []
    for c in CANDIDATES:
        if c["hire_type"] in ("fresher", "campus", "intern"):
            rows.append([c["id"], c["name"], "N/A", "N/A", "N/A", "exempt (fresher/campus/intern)"])
            continue
        prev_end = dplus(c["doj"], -75)
        prev_start = dplus(prev_end, -540)
        rows.append([c["id"], c["name"], "Previous Employer Pvt Ltd", prev_start, prev_end, "clean"])
        if c["uan_overlap"]:
            overlap_start = dplus(prev_end, -47)
            overlap_end = dplus(prev_end, 15)
            rows.append([c["id"], c["name"], "Second Employer Services LLP",
                         overlap_start, overlap_end,
                         "OVERLAP with row above -- dual employment flag"])
    write_csv("offer-desk-inputs/03-system-logs/uan-service-history-sample.csv",
              ["candidate_id", "name", "employer", "start_date", "end_date", "note"], rows)


# --------------------------------------------------------- 03 OneDrive -----
def build_onedrive_log():
    rows = []
    for c in CANDIDATES:
        if c["uan_overlap"]:
            continue
        folder = f"/HR-Ops/{c['doj'][:7]}/{c['doj']}/{c['id']}-{c['name'].replace(' ', '')}/"
        for f in ["Offer_Letter_Signed.pdf", "Onboarding_Ack_Signed.pdf", "Document_Checklist.xlsx"]:
            rows.append([c["id"], folder + f, dplus(c["doj"], -2), "Rashmi KN"])
    write_csv("offer-desk-inputs/03-system-logs/onedrive-placement-log.csv",
              ["candidate_id", "file_path", "placed_date", "placed_by"], rows)


# ------------------------------------------------- 03 Master Joining Sheet -
def build_mjs():
    headers = ["Candidate ID", "Name", "Hire Type", "Function", "Grade", "DOJ", "City",
               "Recruiter", "Manager", "CTC", "Variable %", "Joining Bonus",
               "Deferred Bonus", "ESOP", "Email ID (<=18 char)", "IT Ticket Status",
               "Onboarding Ack Status", "Notes"]
    rows = []
    for c in CANDIDATES:
        first = c["name"].split()[0].lower()
        last = c["name"].split()[-1].lower()
        email = f"{first}.{last}@company.com"
        if len(email.split("@")[0]) > 18:
            email = f"{first[:1]}.{last}@company.com"
        rows.append([
            c["id"], c["name"], c["hire_type"], c["function"], c["grade"], c["doj"], c["city"],
            c["recruiter"], c["manager"], c["ctc"] or "", c["variable_pct"] or "",
            c["joining_bonus"] or "", c["deferred_bonus"] or "", "Yes" if c["esop"] else "No",
            email, "closed" if not c["uan_overlap"] else "on-hold",
            "signed" if not c["uan_overlap"] else "not-sent", c["flag"],
        ])
    write_xlsx("offer-desk-inputs/03-system-logs/master-joining-sheet.xlsx", headers, rows, "MJS")


def build_pr_sheet():
    headers = ["Candidate ID", "Name", "Skills", "Contract Start", "Contract End",
               "Contract Amount (INR)", "Asset Payroll", "Vendor"]
    rows = []
    for c in CANDIDATES:
        if c["hire_type"] != "contractor":
            continue
        rows.append([c["id"], c["name"], "As per SOW", c["doj"],
                     dplus(c["doj"], 30 * c["contract_months"]), c["contract_amount"],
                     "Yes" if c["asset_payroll"] else "No", c["source_name"]])
    write_xlsx("offer-desk-inputs/03-system-logs/pr-sheet-contractors.xlsx", headers, rows, "PR Sheet")


def build_email_tracker():
    headers = ["Candidate ID", "Name", "Proposed Email", "Length OK (<=18)", "IT Ticket Sent",
               "IT Response"]
    rows = []
    for c in CANDIDATES:
        if c["uan_overlap"]:
            continue
        first = c["name"].split()[0].lower()
        last = c["name"].split()[-1].lower()
        email = f"{first}.{last}"
        ok = len(email) <= 18
        if not ok:
            email = f"{first[:1]}.{last}"
        rows.append([c["id"], c["name"], email, "Yes" if len(email) <= 18 else "No",
                     dplus(c["doj"], -1), "confirmed"])
    write_xlsx("offer-desk-inputs/03-system-logs/email-id-creation-tracker.xlsx", headers, rows, "Email IDs")


def build_payroll_report():
    headers = ["Candidate ID", "Name", "Joining Bonus Installment 1", "Installment 1 Month",
               "Joining Bonus Installment 2", "Installment 2 Month", "Deferred Bonus",
               "Deferred Vest Date"]
    rows = []
    for c in CANDIDATES:
        if c["joining_bonus"]:
            half = c["joining_bonus"] / 2 if c["joining_bonus"] > 100000 else c["joining_bonus"]
            m4 = dplus(c["doj"], 120)[:7]
            m8 = dplus(c["doj"], 240)[:7] if c["joining_bonus"] > 100000 else ""
            rows.append([c["id"], c["name"], half, m4,
                         half if c["joining_bonus"] > 100000 else "", m8,
                         c["deferred_bonus"] or "", c["deferred_bonus_vest"]])
    write_xlsx("offer-desk-inputs/03-system-logs/payroll-report-17th.xlsx", headers, rows, "Payroll")


# ------------------------------------------------------------- 04 ----------
def build_throughput_log():
    # 6 months of candidate volume, independent of the 9-candidate roster
    # (roster = 1 illustrative batch; this = the real monthly pattern behind
    # the 61.8-hours defended-case math).
    rows = [
        ("2026-02", 34, 3), ("2026-03", 41, 4), ("2026-04", 29, 2),
        ("2026-05", 38, 3), ("2026-06", 44, 5), ("2026-07", 9, 1),
    ]
    write_csv("offer-desk-inputs/04-measurement/monthly-throughput-log.csv",
              ["month", "offers_processed", "dual_employment_stops"], rows)


# ------------------------------------------------------------- 05 ----------
def build_consent_receipt():
    write_json("offer-desk-inputs/05-governance/consent-receipt-record.json", {
        "_disclaimer": "FABRICATED FOR TESTING. Payload shape matches POST /api/consent/receipts.",
        "interview_ref": "Rashmi KN, Offer Desk SME sitting, 12 May 2026",
        "purpose": "Scout interview capture for the Offer Desk worked example -- HR & People Ops function census",
        "consent_timestamp": "2026-05-12T15:05:00+05:30",
        "retention_days": 90,
        "status": "active",
    })


# ------------------------------------------------------------- 02 ----------
def build_salary_grid():
    headers = ["Function", "Grade", "Min (INR)", "Max (INR)", "Variable Pay %",
               "Deviation Approval Required Above", "Approvers"]
    rows = [
        ["GTS", "Grade 3", 900000, 1300000, 0, "5%", "Vasu, Nagaraj"],
        ["GTS", "Grade 4", 1200000, 1700000, 0, "5%", "Vasu, Nagaraj"],
        ["GTS", "Grade 5", 1600000, 2200000, 8, "5%", "Vasu, Nagaraj"],
        ["GTS", "Grade 6", 2100000, 2900000, 8, "5%", "Vasu, Nagaraj"],
        ["GTS", "Grade 7", 2800000, 3800000, 8, "5%", "Vasu, Nagaraj"],
        ["Platforms", "Grade 3", 950000, 1350000, 0, "5%", "Vasu, Nagaraj"],
        ["Platforms", "Grade 4", 1250000, 1750000, 0, "5%", "Vasu, Nagaraj"],
        ["Platforms", "Grade 5", 1700000, 2300000, 0, "5%", "Vasu, Nagaraj"],
        ["Platforms", "Grade 6", 2200000, 3000000, 0, "5%", "Vasu, Nagaraj"],
        ["Platforms", "Grade 7", 2900000, 3900000, 0, "5%", "Vasu, Nagaraj"],
        ["Core", "Grade 3", 850000, 1250000, 0, "5%", "Vasu, Nagaraj"],
        ["Core", "Grade 4", 1150000, 1600000, 0, "5%", "Vasu, Nagaraj"],
        ["Core", "Grade 5", 1550000, 2100000, 8, "5%", "Vasu, Nagaraj"],
    ]
    write_xlsx("offer-desk-inputs/02-policy/salary-grid-and-deviation-policy.xlsx",
               headers, rows, "Salary Grid")


if __name__ == "__main__":
    build_candidate_roster()
    build_zwayam_export()
    build_zoho_log()
    build_uan_history()
    build_onedrive_log()
    build_mjs()
    build_pr_sheet()
    build_email_tracker()
    build_payroll_report()
    build_throughput_log()
    build_consent_receipt()
    build_salary_grid()
    print("done")
