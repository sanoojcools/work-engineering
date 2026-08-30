"""
Bulk-create Work Units from Excel/CSV via the live API.

This repo has no bulk HTTP endpoint. The UI creates one unit at a time;
this client POSTs each row to /api/work-units/.

Usage (from repo root, backend on :8000):

  py -3.12 -m pip install openpyxl
  py -3.12 bulk_ingest.py --init-template
  py -3.12 bulk_ingest.py --file HR_Work_Units_Bulk.xlsx --api http://localhost:8000

Python 3.14 is fine for this client. --key is unused: Work Unit create is not Spec API.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

METHODS = {
    "deterministic_rule",
    "database_constraint",
    "cross_system_reconciliation",
    "human_spot_check",
    "llm_as_judge",
    "outcome_delay",
    "counterparty_confirmation",
}
METHOD_ALIASES = {
    "deterministic": "deterministic_rule",
    "rule": "deterministic_rule",
    "auto": "deterministic_rule",
    "db": "database_constraint",
    "database": "database_constraint",
    "recon": "cross_system_reconciliation",
    "reconciliation": "cross_system_reconciliation",
    "human": "human_spot_check",
    "spot": "human_spot_check",
    "spot_check": "human_spot_check",
    "llm": "llm_as_judge",
    "judge": "llm_as_judge",
    "delay": "outcome_delay",
    "counterparty": "counterparty_confirmation",
}
ACTORS = {"human", "agent", "deterministic", "external"}
HEADERS = [
    "Code",
    "Title",
    "Business Object",
    "Owner / Authority",
    "Owner Type",
    "Current Condition",
    "Desired Condition",
    "Acceptance Criteria",
    "Verification Method",
    "Evidence Required",
    "Do Time (mins)",
    "Verify Time (mins)",
    "Exception %",
    "Context",
    "Trigger",
    "Inputs",
    "SLA hours",
]
SAMPLE_ROWS = [
    [
        "WU-ONB-01",
        "Collect joining documents",
        "Employee",
        "HR Ops SPOC",
        "human",
        "Offer status = signed in Zoho",
        "Documents status = received",
        "All mandatory docs attached in the HRIS",
        "human_spot_check",
        "HRIS document packet ID",
        "8",
        "4",
        "10",
        "Pre-joining document chase",
        "Offer signed in Zoho",
        "Candidate name, offer id",
        "24",
    ],
    [
        "WU-ONB-04",
        "Pre-Joining Communication",
        "Employee",
        "HR Ops SPOC",
        "agent",
        "Offer status = signed in Zoho",
        "Welcome mail status = delivered",
        "Outlook mail log exists AND Teams invite exists",
        "deterministic_rule",
        "Outlook message ID + Teams invite ID",
        "2",
        "1",
        "5",
        "Pre-joining welcome for a signed offer",
        "Offer signed in Zoho",
        "Candidate name, joining date, offer id",
        "8",
    ],
    [
        "WU-ONB-05",
        "Create HRIS employee record",
        "Employee",
        "HR Ops SPOC",
        "human",
        "Welcome mail status = delivered",
        "Employee status = pre_joining",
        "HRIS employee id exists",
        "database_constraint",
        "HRIS employee id",
        "6",
        "3",
        "8",
        "Master data create after welcome mail",
        "Welcome mail delivered",
        "Candidate name, offer id",
        "8",
    ],
]


def norm(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return text


def clip(value: str, limit: int, label: str, code: str) -> str:
    if len(value) <= limit:
        return value
    print(f"  WARN {code}: {label} truncated to {limit} chars")
    return value[:limit]


def method_of(raw: str) -> str:
    key = raw.strip().lower().replace(" ", "_").replace("-", "_")
    if key in METHODS:
        return key
    return METHOD_ALIASES.get(key, "deterministic_rule")


def actor_of(raw: str) -> str:
    key = raw.strip().lower()
    if key in ACTORS:
        return key
    if "agent" in key or "robot" in key:
        return "agent"
    if "deterministic" in key or "auto" in key:
        return "deterministic"
    if "external" in key or "bpo" in key:
        return "external"
    return "human"


class Api:
    def __init__(self, base: str) -> None:
        self.base = base.rstrip("/")

    def request(self, method: str, path: str, body: dict | None = None) -> tuple[int, object]:
        data = None if body is None else json.dumps(body).encode("utf-8")
        req = urllib.request.Request(
            self.base + path,
            data=data,
            method=method,
            headers={"Content-Type": "application/json", "Accept": "application/json"},
        )
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                raw = resp.read().decode("utf-8")
                return resp.status, json.loads(raw) if raw else {}
        except urllib.error.HTTPError as err:
            raw = err.read().decode("utf-8", errors="replace")
            try:
                parsed: object = json.loads(raw)
            except json.JSONDecodeError:
                parsed = raw
            return err.code, parsed


def load_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            return [{norm(k): norm(v) for k, v in row.items()} for row in csv.DictReader(handle)]
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("xlsx needs openpyxl. Run: python -m pip install openpyxl") from exc
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = [norm(cell) for cell in next(rows_iter)]
    out: list[dict[str, str]] = []
    for values in rows_iter:
        row = {header[i]: norm(values[i] if i < len(values) else "") for i in range(len(header))}
        if any(row.values()):
            out.append(row)
    return out


def col(row: dict[str, str], *names: str, default: str = "") -> str:
    lowered = {key.lower(): value for key, value in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value:
            return value
    return default


def write_template(path: Path) -> None:
    if path.suffix.lower() == ".csv":
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(HEADERS)
            writer.writerows(SAMPLE_ROWS)
        return
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("xlsx needs openpyxl. Run: python -m pip install openpyxl") from exc
    book = Workbook()
    sheet = book.active
    sheet.title = "Work Units"
    sheet.append(HEADERS)
    for row in SAMPLE_ROWS:
        sheet.append(row)
    book.save(path)


def ensure_types(api: Api, names: list[str]) -> dict[str, int]:
    status, payload = api.request("GET", "/api/ontology/types")
    if status != 200 or not isinstance(payload, dict):
        raise SystemExit(f"Cannot list ontology types ({status}): {payload}")
    by_name = {item["name"]: item["id"] for item in payload.get("items", [])}
    for name in names:
        if name in by_name:
            print(f"  Ontology type exists: {name} (id {by_name[name]})")
            continue
        status, created = api.request(
            "POST",
            "/api/ontology/types",
            {
                "name": name,
                "kind": "business_object",
                "description": f"{name} (bulk ingest)",
                "state_machine": '["draft","pre_joining","active","on_hold","offboarded","exited"]',
            },
        )
        if status in (200, 201) and isinstance(created, dict):
            by_name[name] = created["id"]
            print(f"  Created ontology type: {name} (id {created['id']})")
        elif status == 409:
            status, payload = api.request("GET", "/api/ontology/types")
            by_name = {item["name"]: item["id"] for item in payload.get("items", [])}  # type: ignore[union-attr]
            print(f"  Ontology type already existed: {name}")
        else:
            print(f"  FAILED ontology type {name}: {status} {created}")
    return by_name


def existing_units(api: Api, client_id: int | None = None) -> dict[str, dict]:
    path = "/api/work-units/"
    if client_id is not None:
        path += f"?client_id={client_id}"
    status, payload = api.request("GET", path)
    if status != 200 or not isinstance(payload, dict):
        raise SystemExit(f"Cannot list Work Units ({status}): {payload}")
    return {item["code"]: item for item in payload.get("items", [])}


def ingest_row(api: Api, row: dict[str, str], types: dict[str, int], known: dict[str, dict], seed_runs: bool, client_id: int | None = None) -> str:
    code = clip(col(row, "Code", "code", "ID"), 40, "code", "?")
    if not code:
        return "skip"
    if code in known:
        print(f"  SKIP {code} (already exists)")
        return "skip"

    object_name = col(row, "Business Object", "business_object", "Object")
    type_id = types.get(object_name)
    if not type_id:
        print(f"  FAILED {code}: unknown business object '{object_name}'")
        return "fail"

    owner = col(row, "Owner / Authority", "Owner", "Authority", "owner")
    actor = actor_of(col(row, "Owner Type", "Actor Type", "actor_type", default="human"))
    current = clip(col(row, "Current Condition", "current_condition", "Pre-state"), 80, "current_condition", code)
    desired = clip(col(row, "Desired Condition", "desired_condition", "Post-state"), 80, "desired_condition", code)
    name = clip(col(row, "Title", "Name", "name", default=code), 200, "name", code)
    evidence = col(row, "Evidence Required", "Evidence", "evidence_required")
    payload = {
        "code": code,
        "name": name,
        "business_object_type_id": type_id,
        "current_condition": current,
        "desired_condition": desired,
        "context": col(row, "Context", "context", default=f"HR operations · {object_name}"),
        "trigger": col(row, "Trigger", "trigger", default=current),
        "inputs": col(row, "Inputs", "inputs", default=evidence or object_name),
        "authority": owner,
        "actor_constraints": col(row, "Actor Constraints", default=actor),
        "acceptance_criteria": col(row, "Acceptance Criteria", "acceptance_criteria"),
        "evidence_required": evidence,
        "verification_method": method_of(col(row, "Verification Method", "verification_method", default="deterministic_rule")),
        "sla_hours": float(col(row, "SLA hours", "sla_hours", default="8") or 8),
        "failure_semantics": col(row, "Failure Semantics", default="Hold; notify owner; do not silently retry"),
        "provenance": "designed",
        "owner": owner,
        "actor_type": actor,
        "client_id": client_id,
    }
    status, created = api.request("POST", "/api/work-units/", payload)
    if status == 409:
        print(f"  SKIP {code} (already exists)")
        return "skip"
    if status not in (200, 201) or not isinstance(created, dict):
        print(f"  FAILED {code}: {status} {created}")
        return "fail"

    unit_id = created["id"]
    known[code] = created
    api.request("POST", f"/api/work-units/{unit_id}/reconcile")

    do_mins = col(row, "Do Time (mins)", "Do Time", "minutes_per_execution")
    verify_mins = col(row, "Verify Time (mins)", "Verify Time", "verification_minutes")
    exception = col(row, "Exception %", "exception_rate", default="5")
    if do_mins or verify_mins:
        rate = float(exception or 0)
        if rate > 1:
            rate = rate / 100
        api.request(
            "PUT",
            f"/api/economics/{unit_id}",
            {
                "executions_per_month": 50,
                "minutes_per_execution": float(do_mins or 0),
                "verification_minutes": float(verify_mins or 0),
                "failure_rate": min(max(rate, 0), 1),
                "exception_minutes": 20,
                "maintenance_hours": 1,
                "attribution_confidence": 0.8,
            },
        )

    if seed_runs:
        for i in range(1, 6):
            api.request(
                "POST",
                "/api/verification/runs",
                {
                    "work_unit_id": unit_id,
                    "method": payload["verification_method"],
                    "independent": True,
                    "outcome": "passed",
                    "evidence_ref": f"MSG-{100 + i}",
                    "notes": f"Bulk seeded run {i}/5",
                },
            )

    print(f"  CREATED {code} — {name}")
    return "created"


def main() -> int:
    parser = argparse.ArgumentParser(description="Bulk ingest Work Units into the Work Engineering API")
    parser.add_argument("--file", default="HR_Work_Units_Bulk.xlsx")
    parser.add_argument("--api", default="http://localhost:8000")
    parser.add_argument("--key", default="dev-spec-key-change-me", help="Ignored; create is not Spec API")
    parser.add_argument("--init-template", action="store_true", help="Write a starter Excel/CSV and exit")
    parser.add_argument("--client", default="catalog", help="Company slug (catalog or client-a)")
    args = parser.parse_args()

    path = Path(args.file)
    if args.init_template:
        if path.suffix.lower() not in {".xlsx", ".csv"}:
            path = Path("HR_Work_Units_Bulk.xlsx")
        write_template(path)
        print(f"Wrote template {path.resolve()}")
        return 0

    if not path.exists():
        csv_fallback = path.with_suffix(".csv")
        if csv_fallback.exists():
            path = csv_fallback
        else:
            print(f"File not found: {path.resolve()}")
            print("That workbook is not in this repo. Put it here, or write a starter file:")
            print(f"  python bulk_ingest.py --init-template --file {args.file}")
            return 1

    rows = load_rows(path)
    print(f"Loaded {len(rows)} rows from {path}")

    api = Api(args.api)
    health_status, health = api.request("GET", "/api/health")
    if health_status != 200:
        print(f"API not reachable at {args.api} ({health_status}): {health}")
        return 1
    print(f"API: {health}")

    objects = sorted({col(row, "Business Object", "business_object") for row in rows if col(row, "Business Object", "business_object")})
    types = ensure_types(api, objects)
    client_id = None
    cl_status, cl_payload = api.request("GET", "/api/clients/")
    if cl_status == 200 and isinstance(cl_payload, dict):
        match = next((c for c in cl_payload.get("items", []) if c.get("slug") == args.client), None)
        if match:
            client_id = match["id"]
            print(f"Company: {match.get('name')} (id {client_id})")
    known = existing_units(api, client_id)

    counts = {"created": 0, "skip": 0, "fail": 0}
    for row in rows:
        result = ingest_row(api, row, types, known, args.seed_runs, client_id)
        counts[result] += 1

    print(f"\nDone. created={counts['created']} skipped={counts['skip']} failed={counts['fail']}")
    print("Refresh Work Units in the UI.")
    return 0 if counts["fail"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
