"""Slice 1 PR 1b (playbook E.2): format classifier + review queue.

Detects the header row of an uploaded CSV/XLSX by scoring each cell against
the Step-identity synonym dict. This is deliberately dumb and deterministic
— no LLM anywhere in this module. If nothing in the file scores above the
queue threshold, the WHOLE FILE goes to review_queue and nothing is ever
handed to the import service. That's the point: a File-6-class sheet must
never become a hallucinated genome.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

# Confidence rule — published, do not invent another (playbook E.0).
CONFIDENCE_EXACT = 1.0
CONFIDENCE_SYNONYM = 0.8
CONFIDENCE_FUZZY = 0.6
CONFIDENCE_ELSE = 0.3
CONFIDENCE_QUEUE_THRESHOLD = 0.7  # < this -> review_queue, not import

# Synonym seed for the Step identity column (playbook E.0) — extend only with tests.
STEP_IDENTITY_CANONICAL = "Step #"
STEP_IDENTITY_SYNONYMS = ["Step No", "Sl No", "Activity", "Task", "S.No", "No"]
STEP_IDENTITY_ALL = [STEP_IDENTITY_CANONICAL, *STEP_IDENTITY_SYNONYMS]

METADATA_HINT_LABELS = {"workflow name", "outcome", "trigger", "frequency", "spoc", "category"}
MAX_HEADER_SCAN_ROWS = 50


def levenshtein(a: str, b: str) -> int:
    """Classic edit distance, pure Python, no dependency — case/whitespace
    insensitive since header text varies (" Step No" vs "step no")."""
    a, b = a.strip().lower(), b.strip().lower()
    if a == b:
        return 0
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        curr = [i] + [0] * len(b)
        for j, cb in enumerate(b, start=1):
            cost = 0 if ca == cb else 1
            curr[j] = min(curr[j - 1] + 1, prev[j] + 1, prev[j - 1] + cost)
        prev = curr
    return prev[-1]


def score_header_cell(cell: str) -> float:
    h = (cell or "").strip()
    if not h:
        return CONFIDENCE_ELSE
    if h == STEP_IDENTITY_CANONICAL:
        return CONFIDENCE_EXACT
    if h in STEP_IDENTITY_SYNONYMS:
        return CONFIDENCE_SYNONYM
    best = min(levenshtein(h, s) for s in STEP_IDENTITY_ALL)
    if best <= 2:
        return CONFIDENCE_FUZZY
    return CONFIDENCE_ELSE


@dataclass
class ClassificationResult:
    queued: bool
    header_row_index: int | None = None
    step_identity_column: int | None = None
    step_identity_confidence: float = 0.0
    header_cells: list[str] = field(default_factory=list)
    metadata_notes: dict[str, str] = field(default_factory=dict)
    raw_header_row_text: str = ""


def _rows_from_bytes(content: bytes, file_name: str) -> list[list[str]]:
    if file_name.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def classify(content: bytes, file_name: str) -> ClassificationResult:
    """Scans rows top-to-bottom. A row whose first cell matches a known
    metadata-block label (Workflow Name, Outcome, ...) is captured as a
    file-level note, not a header candidate — playbook: "do not invent WUs
    from them." The header row is the first row with any cell scoring
    >= 0.7. If none is found in the scan window, the whole file queues."""
    rows = _rows_from_bytes(content, file_name)
    metadata_notes: dict[str, str] = {}

    for row_idx, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        if not any((c or "").strip() for c in row):
            continue

        label = (row[0] or "").strip().lower()
        if label in METADATA_HINT_LABELS and len(row) > 1 and (row[1] or "").strip():
            metadata_notes[row[0].strip()] = row[1].strip()
            continue

        scores = [score_header_cell(c) for c in row]
        best_score = max(scores) if scores else 0.0
        if best_score >= CONFIDENCE_QUEUE_THRESHOLD:
            best_col = scores.index(best_score)
            return ClassificationResult(
                queued=False,
                header_row_index=row_idx,
                step_identity_column=best_col,
                step_identity_confidence=best_score,
                header_cells=[str(c) for c in row],
                metadata_notes=metadata_notes,
                raw_header_row_text=",".join(str(c) for c in row),
            )

    return ClassificationResult(queued=True, metadata_notes=metadata_notes)
